import torch
import numpy as np
import pandas as pd
import networkx as nx
from texttable import Texttable
from scipy.sparse import coo_matrix
from torch_geometric.datasets import PPI
from torch_geometric.datasets import Planetoid
from torch_geometric.datasets import WikiCS
from torch_geometric.utils.convert import to_networkx

def tab_printer(args):
    """
    Function to print the logs in a nice tabular format.
    :param args: Parameters used for the model.
    """
    args = vars(args)
    keys = sorted(args.keys())
    t = Texttable() 
    t.add_rows([["Parameter", "Value"]] +  [[k.replace("_"," ").capitalize(),args[k]] for k in keys])
    print(t.draw())

def graph_reader(path):
    """
    Function to read the graph from the path.
    :param path: Path to the edge list.
    :return graph: NetworkX object returned.
    """
    graph = nx.from_edgelist(pd.read_csv(path).values.tolist())
    return graph

def feature_reader(path):
    """
    Reading the sparse feature matrix stored as csv from the disk.
    :param path: Path to the csv file.
    :return features: Dense matrix of features.
    """
    features = pd.read_csv(path)
    node_index = features["node_id"].values.tolist()
    feature_index = features["feature_id"].values.tolist()
    feature_values = features["value"].values.tolist()
    node_count = max(node_index)+1
    feature_count = max(feature_index)+1
    features = coo_matrix((feature_values, (node_index, feature_index)), shape=(node_count, feature_count)).toarray()
    return features
    # Number of graph nodes:  19717
    # (19717, 500) (19717, 1)

def target_reader(path):
    """
    Reading the target vector from disk.
    :param path: Path to the target.
    :return target: Target vector.
    """
    target = np.array(pd.read_csv(path)["target"]).reshape(-1,1)
    return target

def dataset_reader(args):
    """
    Reading the dataset
    :param dataset_name: Name of the dataset.
    :param path: Path to the target.
    :return target: Target vector.
    """
    dataset_name = args.dataset_name
    # if dataset_name=='PPI':        
    #     dataset = PPI(root=path)
    if dataset_name=='default':
        graph = graph_reader("./input/edges.csv")
        features = feature_reader("./input/features.csv")
        target = target_reader("./input/target.csv")

    elif dataset_name in ['PubMed', 'Cora', 'CiteSeer','WikiCS']:
        if dataset_name == 'PubMed':
            dataset = Planetoid(root=args.ds_root+'/PubMed', name='PubMed', split='full')
        elif dataset_name == 'Cora':
            dataset = Planetoid(root=args.ds_root+'/Cora', name='Cora', split='full')
        elif dataset_name == 'CiteSeer':
            dataset = Planetoid(root=args.ds_root+'/CiteSeer', name='CiteSeer', split='full')
        elif dataset_name == 'WikiCS':
            dataset = WikiCS(root=args.ds_root+'/WikiCS')
        data = dataset[0]
        graph = to_networkx(data, to_undirected=True)
        # node_labels = data.y[list(graph.nodes)].numpy()
        # len(graph.nodes()), len(graph.edges())
        features = data.x.numpy()
        target = data.y.numpy()[..., np.newaxis]

    return graph, features, target


import os
from openpyxl import load_workbook


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()